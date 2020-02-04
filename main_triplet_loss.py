# %%
# TODO: This makes sure the weights are initialized the same, but results after training still not reproducible
seed_value = 1
import os
os.environ['PYTHONHASHSEED']=str(seed_value)
os.environ['TF_CUDNN_DETERMINISTIC'] = '1'  # new flag present in tf 2.0+
import random
random.seed(seed_value)
import numpy as np 
np.random.seed(seed_value)
import tensorflow as tf 
tf.random.set_seed(seed_value) 

import DLCVDatasets
import models
import utils
import matplotlib.pyplot as plt
import train_triplet_loss
import os
import csv

# %% Load and preprocess data
dataset_name = 'mnist'    # mnist or cifar10
train_size = 60000
test_size = 10000
used_labels = [0,1,2,3,4,5,6,7,8]    # the labels to be loaded
num_classes = len(used_labels)
x_train, y_train, x_test, y_test, class_names = DLCVDatasets.get_dataset(dataset_name,
                                                                         used_labels=used_labels,
                                                                         training_size=train_size,
                                                                         test_size=test_size)
# Normalization
#mean = np.mean(x_train)
#x_train -= mean
#x_test -= mean
x_train, x_test = x_train / 255.0, x_test / 255.0

# Reshape to add the channel dimension
x_train = np.reshape(x_train, x_train.shape+(1,))
x_test = np.reshape(x_test, x_test.shape+(1,))
input_shape = x_train.shape[1:]

# %% Get the model
encoding_dim = 64
normalized_encodings = False
model = models.get_model_v4(input_shape, encoding_dim, normalized_encodings)
model.summary()

# %% Train the model with triplet loss
num_epochs=6
batch_size = 64
learning_rate = 0.0005
margin=0.5
triplet_loss_strategy="batch_hard"
train_triplet_loss.train_model_with_tripletloss(model, x_train, y_train,
                                              x_test, y_test, num_classes, encoding_dim,
                                              num_epochs, batch_size, learning_rate,
                                              margin,  triplet_loss_strategy)
#model.save('model_trained_TL_lenet_1.h5')

# %% Evaluate the model
# Load the complete dataset, including 0 - 9
used_labels = list(range(0, 10))    # the labels to be loaded
x_train, y_train, x_test, y_test, class_names = DLCVDatasets.get_dataset(dataset_name,
                                                                         used_labels=used_labels,
                                                                         training_size=train_size,
                                                                         test_size=test_size)
# Reshape to add the channel dimension
x_train = np.reshape(x_train, x_train.shape+(1,))
x_test = np.reshape(x_test, x_test.shape+(1,))

# Sort test samples in ascending order (from 0 to 9)
sorted_idc = np.argsort(y_test, kind='stable')
x_test = x_test[sorted_idc]
y_test = y_test[sorted_idc]

# Normalized sorted test data for prediction
x_test_normalized = x_test / 255.0

# Store the indices for each digit in dict
digits_idc = {}
for idx in range(10):
    digits_idc[str(idx)] = np.argwhere(y_test == idx)

# Compute encodings and pairwise euclidean distances
encodings = model.predict(x_test_normalized)
#normalized_encodings = utils.l2_normalize(encodings)
encoding_new = encodings.flatten(order='C')
# plot3D(encoding_new,y_test)
#pairwise_dists = utils.cal_pairwise_dists(normalized_encodings)
pairwise_dists = utils.cal_pairwise_dists(encodings)

# %% Save results for embedding projector

dirname = 'log/tripleLoss_all'
if not os.path.exists(dirname):
    os.makedirs(dirname)
with open(dirname+'/metadata.tsv', 'w') as metadata_file:
    metadata_file.write('Index\tLabel\n')
    for idx, row in enumerate(y_test):
        metadata_file.write('%d\t%d\n' % (idx, row))

with open(dirname+'/feature_vecs.tsv', 'w') as fw:
    csv_writer = csv.writer(fw, delimiter='\t')
    for vec in encodings:
        csv_writer.writerow(vec)

# %%
img2_0_idx = digits_idc['2'][0]
img2_1_idx = digits_idc['2'][20]
img5_0_idx = digits_idc['5'][0]
img5_1_idx = digits_idc['5'][20]
img6_0_idx = digits_idc['6'][0]
img6_1_idx = digits_idc['6'][20]
img9_0_idx = digits_idc['9'][0]
img9_1_idx = digits_idc['9'][20]

encoding_2_0 = encodings[img2_0_idx]
encoding_2_1 = encodings[img2_1_idx]

encoding_5_0 = encodings[img5_0_idx]
encoding_5_1 = encodings[img5_1_idx]

encoding_6_0 = encodings[img6_0_idx]
encoding_6_1 = encodings[img6_1_idx]

encoding_9_0 = encodings[img9_0_idx]
encoding_9_1 = encodings[img9_1_idx]

# Visualization
plt.figure(1)
pixels = np.array(x_test[img2_0_idx], dtype='uint8')
pixels = pixels.reshape((28, 28))
plt.subplot(421)
plt.imshow(pixels, cmap='gray')

pixels = np.array(x_test[img2_1_idx], dtype='uint8')
pixels = pixels.reshape((28, 28))
plt.subplot(422)
plt.imshow(pixels, cmap='gray')

pixels = np.array(x_test[img5_0_idx], dtype='uint8')
pixels = pixels.reshape((28, 28))
plt.subplot(423)
plt.imshow(pixels, cmap='gray')

pixels = np.array(x_test[img5_1_idx], dtype='uint8')
pixels = pixels.reshape((28, 28))
plt.subplot(424)
plt.imshow(pixels, cmap='gray')

pixels = np.array(x_test[img6_0_idx], dtype='uint8')
pixels = pixels.reshape((28, 28))
plt.subplot(425)
plt.imshow(pixels, cmap='gray')

pixels = np.array(x_test[img6_1_idx], dtype='uint8')
pixels = pixels.reshape((28, 28))
plt.subplot(426)
plt.imshow(pixels, cmap='gray')

pixels = np.array(x_test[img9_0_idx], dtype='uint8')
pixels = pixels.reshape((28, 28))
plt.subplot(427)
plt.imshow(pixels, cmap='gray')

pixels = np.array(x_test[img9_1_idx], dtype='uint8')
pixels = pixels.reshape((28, 28))
plt.subplot(428)
plt.imshow(pixels, cmap='gray')
plt.show()

print('2 & 2: {}'.format(tf.norm(encoding_2_0 - encoding_2_1).numpy()))
print('2 & 5: {}'.format(tf.norm(encoding_2_0 - encoding_5_0).numpy()))
print('2 & 6: {}'.format(tf.norm(encoding_2_0 - encoding_6_0).numpy()))

print('5 & 5: {}'.format(tf.norm(encoding_5_0 - encoding_5_1).numpy()))
print('6 & 6: {}'.format(tf.norm(encoding_6_0 - encoding_6_1).numpy()))
print('5 & 6: {}'.format(tf.norm(encoding_5_0 - encoding_6_0).numpy()))

print('9 & 9: {}'.format(tf.norm(encoding_9_0 - encoding_9_1).numpy()))
print('5 & 9: {}'.format(tf.norm(encoding_5_0 - encoding_9_0).numpy()))
print('6 & 9: {}'.format(tf.norm(encoding_6_0 - encoding_9_0).numpy()))

# %% Performance evaluation

utils.threshold_evaluation(pairwise_dists, y_test, 2.5, 3.5, 11, i_want_to_plot = True)

treshold = 2.7
(recall, FAR, precision) = utils.performance_test(pairwise_dists, y_test, treshold)
accuracy_table = utils.get_accuracy_table(pairwise_dists, y_test, treshold )

# %% save accuracy table in Excell
import openpyxl
wb =  openpyxl.Workbook()
ws =  wb.active
for i in range(10):
    ws.cell(row=i+2, column=1).value = i
    ws.cell(row=1, column=i+2).value = i

for (i, j), value in np.ndenumerate(accuracy_table):
     ws.cell(row=i+2, column=j+2).value = value
wb.save('log/tripletloss/accuracy_without_9.xlsx')

# %%
